"""
prerender.py — Build-time score computation.
Reads master.xlsx + every player Excel, scores all groups and subsequent rounds,
writes data/scores.json sorted by total points descending.

Replaces update_manifest.py (also regenerates the player list).
"""
import json, pathlib, re, sys
from datetime import timezone, timedelta

try:
    import openpyxl
except ImportError:
    sys.exit("Missing dependency: pip install openpyxl")

ROOT      = pathlib.Path(__file__).parent
PLAYERS   = ROOT / "data" / "players"
MASTER    = ROOT / "data" / "master.xlsx"
OUT       = ROOT / "data" / "scores.json"
NICKNAMES = ROOT / "data" / "nicknames.json"

GROUPS           = list("ABCDEFGHIJKL")
BONUS_PER_POS    = 1   # must match js/scorer.js

# openpyxl uses 1-based column numbers
COL_A  = 1
COL_X  = 24  # match datetime (naive, treated as UTC-4 = ET/VET)
COL_AA = 27   # home team name
COL_AC = 29   # home goals
COL_AD = 30   # away goals
COL_AF = 32   # away team name
COL_BD = 56   # standings team name

TZ_VET = timezone(timedelta(hours=-4))  # Venezuela/Eastern — same offset

PLACEHOLDER_NAMES = {"nombre", "name", "player", "jugador"}

# ── Rounds configuration ──────────────────────────────────────────────────────
ROUNDS_CONFIG = {
    "groups": {
        "name": "Fase de Grupos",
        "sheet_name": "WORLDCUP",
        "matches": [((4 + 8 * i) + j) for i in range(len(GROUPS)) for j in range(6)],
        "point_rules": { "exact": 6, "one_goal": 4, "outcome": 3, "wrong_one_goal": 1, "nothing": 0 }
    },
    "round_32_16": {
        "name": "16vos y 8vos",
        "sheet_name": "16",
        "matches": list(range(101, 117)) + list(range(120, 128)),
        "point_rules": { "exact": 10, "one_goal": 8, "outcome": 6, "wrong_one_goal": 2, "nothing": 0 }
    },
    "quarters": {
        "name": "Cuartos",
        "sheet_name": "4",
        "matches": list(range(131, 135)),
        "point_rules": { "exact": 16, "one_goal": 12, "outcome": 8, "wrong_one_goal": 4, "nothing": 0 }
    },
    "semis_3rd": {
        "name": "Semis y 3er puesto",
        "sheet_name": "2",
        "matches": [138, 139, 143],
        "point_rules": { "exact": 24, "one_goal": 18, "outcome": 12, "wrong_one_goal": 6, "nothing": 0 }
    },
    "final": {
        "name": "Final",
        "sheet_name": "1",
        "matches": [147],
        "point_rules": { "exact": 36, "one_goal": 28, "outcome": 18, "wrong_one_goal": 10, "nothing": 0 }
    }
}

# ── Excel parsing ─────────────────────────────────────────────────────────────

def _cell(ws, row, col):
    v = ws.cell(row, col).value
    return None if v == "" or v is None else v

def parse_player_name(wb, fallback):
    """Read Home!C10 for the player's real name; fall back to filename-derived name."""
    home = wb["Home"] if "Home" in wb.sheetnames else None
    if home:
        v = home.cell(10, 3).value  # C10
        if v and str(v).strip() and str(v).strip().lower() not in PLACEHOLDER_NAMES:
            return str(v).strip()
    return fallback

def _dt_to_ts(dt):
    """Convert a naive datetime (treated as VET/ET UTC-4) to a Unix timestamp."""
    if dt is None:
        return None
    try:
        return int(dt.replace(tzinfo=TZ_VET).timestamp())
    except Exception:
        return None

def extract_fixtures_from_master(wb):
    """Extract all 104 match records from the master workbook."""
    ws = wb["WORLDCUP"]
    records = []
    for r_id, cfg in ROUNDS_CONFIG.items():
        for r in cfg["matches"]:
            dt = ws.cell(r, COL_X).value
            records.append({
                "ts":        _dt_to_ts(dt),
                "home":      str(_cell(ws, r, COL_AA) or "?"),
                "away":      str(_cell(ws, r, COL_AF) or "?"),
                "homeGoals": int(v) if (v := _cell(ws, r, COL_AC)) is not None else None,
                "awayGoals": int(v) if (v := _cell(ws, r, COL_AD)) is not None else None,
            })
    return records

def extract_predictions_flat(wb):
    """Extract predicted goals for all 104 matches from a player workbook.

    If a knockout sheet is not present, we return empty predictions [None, None].
    """
    out = []
    for r_id, cfg in ROUNDS_CONFIG.items():
        if r_id == "groups":
            ws = wb["WORLDCUP"]
        else:
            sh_name = cfg["sheet_name"]
            ws = wb[sh_name] if sh_name in wb.sheetnames else None

        for r in cfg["matches"]:
            if ws is None:
                out.append([None, None])
            else:
                hg = _cell(ws, r, COL_AC)
                ag = _cell(ws, r, COL_AD)
                out.append([
                    int(hg) if hg is not None else None,
                    int(ag) if ag is not None else None,
                ])
    return out

# Grid display name overrides — when the first name of the displayName
# doesn't match the real name (e.g. nickname IS the name, not in quotes).
_GRID_OVERRIDES = {
    'Serginho el m\u00edtico': 'Sergio',
    'Emmanuel Lambaz':      'Emma',
}

def grid_short_name(display_name, all_display_names):
    """Return the shortest unambiguous first name for the fixture grid."""
    if display_name in _GRID_OVERRIDES:
        return _GRID_OVERRIDES[display_name]
    clean = re.sub(r'"[^"]*"', '', display_name).strip()
    parts = clean.split()
    first = parts[0] if parts else display_name

    # Check for collisions with any other display name's first word
    others = [
        re.sub(r'"[^"]*"', '', n).strip().split()[0]
        for n in all_display_names
        if n != display_name
    ]
    if first in others and len(parts) >= 2:
        return f"{first} {parts[-1][0]}."
    return first

def parse_groups_from_wb(wb):
    ws = wb["WORLDCUP"]
    groups = []
    for i, letter in enumerate(GROUPS):
        match_start = 4 + 8 * i
        team_start  = 5 + 8 * i

        teams = []
        for j in range(4):
            v = _cell(ws, team_start + j, COL_A)
            teams.append(str(v) if v else f"Equipo${j+1}")

        matches = []
        for j in range(6):
            r = match_start + j
            home = str(_cell(ws, r, COL_AA) or "?")
            away = str(_cell(ws, r, COL_AF) or "?")
            hg   = _cell(ws, r, COL_AC)
            ag   = _cell(ws, r, COL_AD)
            matches.append({
                "home": home, "away": away,
                "homeGoals": int(hg) if hg is not None else None,
                "awayGoals": int(ag) if ag is not None else None,
            })

        groups.append({"letter": letter, "teams": teams, "matches": matches})

    player_name = parse_player_name(wb, "Jugador")
    return player_name, groups, wb

def parse_groups(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    if "WORLDCUP" not in wb.sheetnames:
        raise ValueError(f"No WORLDCUP sheet in {path.name}")
    player_name, groups, wb = parse_groups_from_wb(wb)
    player_name = parse_player_name(wb, display_name(path.name))
    return player_name, groups, wb

# ── Scoring (mirrors js/scorer.js) ───────────────────────────────────────────

def _outcome(h, a):
    return 1 if h > a else (-1 if h < a else 0)

def score_match_for_rules(ph, pa, rh, ra, point_rules):
    if rh is None or ra is None: return None   # not played
    if ph is None or pa is None: return 0, "nothing"      # no prediction
    exact_h = ph == rh
    exact_a = pa == ra
    correct  = _outcome(ph, pa) == _outcome(rh, ra)
    one_goal = exact_h or exact_a
    if exact_h and exact_a:  return point_rules["exact"], "exact"
    if correct and one_goal: return point_rules["one_goal"], "one_goal"
    if correct:              return point_rules["outcome"], "outcome"
    if one_goal:             return point_rules["wrong_one_goal"], "wrong_one_goal"
    return point_rules["nothing"], "nothing"

def compute_standings(teams, matches):
    stats = {t: {"pts": 0, "gf": 0, "ga": 0} for t in teams}
    for m in matches:
        hg, ag = m["homeGoals"], m["awayGoals"]
        if hg is None or ag is None: continue
        h, a = m["home"], m["away"]
        stats[h]["gf"] += hg; stats[h]["ga"] += ag
        stats[a]["gf"] += ag; stats[a]["ga"] += hg
        if hg > ag:   stats[h]["pts"] += 3
        elif hg < ag: stats[a]["pts"] += 3
        else:         stats[h]["pts"] += 1; stats[a]["pts"] += 1
    return sorted(teams, key=lambda t: (
        -stats[t]["pts"],
        -(stats[t]["gf"] - stats[t]["ga"]),
        -stats[t]["gf"],
        t,
    ))

def score_player_all_rounds(player_wb, master_wb, master_groups):
    """Score a player for all rounds and compile the points and counts."""
    rounds_data = {}
    overall_total = 0

    for r_id, cfg in ROUNDS_CONFIG.items():
        if r_id == "groups":
            p_ws = player_wb["WORLDCUP"]
        else:
            sh_name = cfg["sheet_name"]
            if sh_name in player_wb.sheetnames:
                p_ws = player_wb[sh_name]
            else:
                p_ws = None

        m_ws = master_wb["WORLDCUP"]

        points = 0
        counts = { "exact": 0, "one_goal": 0, "outcome": 0, "wrong_one_goal": 0, "nothing": 0 }

        if r_id != "groups" and p_ws is None:
            rounds_data[r_id] = {
                "points": 0,
                "counts": counts,
                "bonus": 0
            }
            continue

        # Score individual matches in this round
        for r in cfg["matches"]:
            ph = _cell(p_ws, r, COL_AC)
            pa = _cell(p_ws, r, COL_AD)
            rh = _cell(m_ws, r, COL_AC)
            ra = _cell(m_ws, r, COL_AD)

            ph = int(ph) if ph is not None else None
            pa = int(pa) if pa is not None else None
            rh = int(rh) if rh is not None else None
            ra = int(ra) if ra is not None else None

            res = score_match_for_rules(ph, pa, rh, ra, cfg["point_rules"])
            if res is not None:
                pts, tier = res
                points += pts
                counts[tier] += 1

        # Standings bonus ONLY for groups
        bonus_points = 0
        if r_id == "groups" and master_groups:
            _, player_groups, _ = parse_groups_from_wb(player_wb)
            for pg, mg in zip(player_groups, master_groups):
                group_done = all(m["homeGoals"] is not None for m in mg["matches"])
                if group_done:
                    real_s = compute_standings(mg["teams"], mg["matches"])
                    pred_s = compute_standings(pg["teams"], pg["matches"])
                    for pos in range(4):
                        if pred_s[pos] == real_s[pos]:
                            bonus_points += BONUS_PER_POS

        rounds_data[r_id] = {
            "points": points + bonus_points,
            "counts": counts,
            "bonus": bonus_points if r_id == "groups" else 0
        }
        overall_total += (points + bonus_points)

    return overall_total, rounds_data

# ── Main ──────────────────────────────────────────────────────────────────────

def display_name(filename):
    base = filename.replace(".xlsx", "").replace(".XLSX", "")
    return base[0].upper() + base[1:] if base else filename

def load_nicknames():
    NICKNAMES_PATH = pathlib.Path("data/nicknames.json")
    if not NICKNAMES_PATH.exists():
        return {}
    return json.loads(NICKNAMES_PATH.read_text(encoding="utf-8"))

def main():
    nicknames = load_nicknames()
    
    # Load master results
    master_wb = None
    master_groups = None
    master_fixtures = []
    if MASTER.exists():
        try:
            master_wb = openpyxl.load_workbook(MASTER, data_only=True)
            _, master_groups, _ = parse_groups(MASTER)
            master_fixtures = extract_fixtures_from_master(master_wb)
            print("master.xlsx cargado")
        except Exception as e:
            print(f" master.xlsx no se pudo leer: {e}")
    else:
        print(" master.xlsx no encontrado — puntos quedarán pendientes")

    # Score every player file
    xlsx_files = sorted(
        f for f in PLAYERS.iterdir()
        if f.suffix.lower() == ".xlsx" and not f.name.startswith("~")
    )

    players = []
    player_preds = {}  # file -> flat predictions list (104 entries)
    for path in xlsx_files:
        fallback = display_name(path.name)
        try:
            player_wb = openpyxl.load_workbook(path, data_only=True)
            excel_name = parse_player_name(player_wb, fallback)
            name = nicknames.get(path.name, excel_name)
            player_preds[path.name] = extract_predictions_flat(player_wb)
            if master_wb:
                pts, rounds_data = score_player_all_rounds(player_wb, master_wb, master_groups)
                print(f"  {name}: {pts} pts")
            else:
                pts, rounds_data = None, None
                print(f"  {name}: pendiente")
        except Exception as e:
            name = nicknames.get(path.name, fallback)
            print(f"    {name}: error — {e}")
            pts, rounds_data = None, None
            player_preds[path.name] = [[None, None]] * 104

        # backward compatibility fallback for old counts structure
        p_counts = None
        if rounds_data and "groups" in rounds_data:
            g_c = rounds_data["groups"]["counts"]
            p_counts = {
                "p6": g_c["exact"],
                "p4": g_c["one_goal"],
                "p3": g_c["outcome"],
                "p1": g_c["wrong_one_goal"],
                "p0": g_c["nothing"]
            }

        players.append({
            "file": path.name,
            "displayName": name,
            "totalPoints": pts,
            "counts": p_counts,
            "rounds": rounds_data
        })

    # Sort primarily by overall total points descending
    def sort_key(p):
        return (
            p["totalPoints"] is None,
            -(p["totalPoints"] or 0),
            p["displayName"].lower(),
        )
    players.sort(key=sort_key)

    # Count total matches played from master
    matches_played = 0
    if master_wb:
        ws = master_wb["WORLDCUP"]
        for r_id, cfg in ROUNDS_CONFIG.items():
            for r in cfg["matches"]:
                if _cell(ws, r, COL_AC) is not None and _cell(ws, r, COL_AD) is not None:
                    matches_played += 1
    print(f"Matches played: {matches_played} / 104")

    # Position history
    history_map  = {}
    prev_matches = None
    if OUT.exists():
        try:
            old = json.loads(OUT.read_text(encoding="utf-8"))
            prev_matches = old.get("matchesPlayed")
            for p in old.get("players", []):
                hist = p.get("positionHistory")
                if not hist and p.get("position") is not None:
                    prev = p.get("prevPosition")
                    hist = ([prev, p["position"]] if prev and prev != p["position"]
                            else [p["position"]])
                history_map[p["file"]] = hist or []
        except Exception:
            pass

    # Assign current rank based on overall total points
    for rank, p in enumerate(players, start=1):
        p["position"] = rank

    new_results = (prev_matches is not None
                   and matches_played > 0
                   and matches_played != prev_matches)

    for p in players:
        hist = list(history_map.get(p["file"]) or [])
        if new_results:
            hist.append(p["position"])
        p["positionHistory"] = hist
        p["prevPosition"]    = hist[-2] if len(hist) >= 2 else None

    # Build fixture grid
    all_names = [p["displayName"] for p in players]
    grid_players = sorted(
        players,
        key=lambda p: grid_short_name(p["displayName"], all_names).lower()
    )
    fixture_grid = [grid_short_name(p["displayName"], all_names) for p in grid_players]

    # Attach predictions to each fixture
    fixtures = []
    for idx, fix in enumerate(master_fixtures):
        preds = []
        for p in grid_players:
            pair = player_preds.get(p["file"], [[None, None]] * 104)
            entry = pair[idx] if idx < len(pair) else [None, None]
            preds.append(entry)
        fixtures.append({
            "ts":        fix["ts"],
            "home":      fix["home"],
            "away":      fix["away"],
            "homeGoals": fix["homeGoals"],
            "awayGoals": fix["awayGoals"],
            "preds":     preds,
        })
    fixtures.sort(key=lambda f: f["ts"] or 0)

    OUT.write_text(json.dumps(
        {
            "matchesPlayed": matches_played,
            "fixtureGrid":   fixture_grid,
            "fixtures":      fixtures,
            "players":       players,
        },
        ensure_ascii=False, indent=2,
    ))
    print(f"scores.json written -- {len(players)} participant(s)")
    if not players:
        sys.exit("ERROR: no player files found -- aborting build")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        sys.exit(f"ERROR: prerender.py crashed: {e}")
